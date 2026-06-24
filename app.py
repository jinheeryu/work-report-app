import streamlit as st
from streamlit_gsheets import GSheetsConnection
import pandas as pd
import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 0. 웹 앱 기본 설정
st.set_page_config(page_title="팀 종합 업무보고 시스템", layout="wide")
st.title("👥 팀 종합 일일 업무 및 계획 관리 시스템")

# 1. 구글 스프레드시트 실시간 연결
try:
    conn = st.connection("gsheets", type=GSheetsConnection)
    df = conn.read(ttl="0d")
except Exception as e:
    st.error("구글 시트 연결에 실패했습니다. Secrets 설정 및 시트 공유 권한을 확인해주세요.")
    df = pd.DataFrame(columns=["날짜", "작성자", "직급", "업무구분", "고객사_프로젝트명", "시작시간", "종료시간", "업무량/내용", "진행상황"])

# 과거 기록 자동완성용 프로젝트 리스트 안전하게 추출
existing_projects = []
if not df.empty and "고객사_프로젝트명" in df.columns:
    existing_projects = sorted(df["고객사_프로젝트명"].dropna().unique().tolist())

col1, col2 = st.columns([1, 1.2])

# ----------------------------------------------------
# 왼쪽 화면: 업무 및 계획 입력 폼 (절대 수정 없음)
# ----------------------------------------------------
with col1:
    st.header("📥 업무 내역 입력")
    
    with st.form("report_form", clear_on_submit=False):
        sub_col1, sub_col2 = st.columns(2)
        with sub_col1:
            author_name = st.text_input("작성자 성명", placeholder="예: 홍길동")
        with sub_col2:
            rank = st.selectbox("직급", ["사원", "대리", "과장", "차장", "부장", "팀장", "기타"])
            
        date_input = st.date_input("업무/계획 날짜 선택", datetime.date.today())
        
        # ⚡ 날짜 기준 오늘 업무 / 내일 계획 자동 판별
        today = datetime.date.today()
        auto_task_type = "내일 계획" if date_input > today else "오늘 업무"
        if auto_task_type == "내일 계획":
            st.info(f"💡 선택한 날짜가 미래이므로 자동으로 **[내일 계획]**으로 분류됩니다.")
        else:
            st.success(f"💡 선택한 날짜가 오늘/과거이므로 자동으로 **[오늘 업무]**로 분류됩니다.")
            
        st.write("---")
        
        # 🔗 프로젝트명 입력 방식
        st.write("🔗 **고객사_프로젝트명 입력**")
        if existing_projects:
            selected_hint = st.selectbox(
                "💡 과거에 입력된 프로젝트 리스트 (참고용)",
                options=["-- 새로 직접 입력하기 --"] + existing_projects,
                index=0
            )
            default_project_text = "" if selected_hint == "-- 새로 직접 입력하기 --" else selected_hint
        else:
            default_project_text = ""
            
        project_name = st.text_input(
            "✍️ 실제 등록할 고객사_프로젝트명 (필수)",
            value=default_project_text,
            placeholder="예: Redhat_마케팅 대시보드 구축"
        )
        
        st.write("---")
        
        # ⏰ 작업 시간 선택 범위 제한 (09:00 ~ 18:00 범위, 15분 단위)
        st.write("⏰ **작업 시간 선택 (09:00 ~ 18:00 범위, 15분 단위)**")
        time_options = []
        for h in range(9, 19):
            for m in (0, 15, 30, 45):
                if h == 18 and m > 0:
                    continue
                time_options.append(f"{h:02d}:{m:02d}")
        
        time_col1, time_col2 = st.columns(2)
        with time_col1:
            start_time_str = st.selectbox("작업 시작 시간", time_options, index=0)
        with time_col2:
            end_time_str = st.selectbox("작업 종료 시간", time_options, index=len(time_options)-1)
            
        description = st.text_area("업무량/내용 상세 기록")
        
        if auto_task_type == "오늘 업무":
            status = st.selectbox("📊 현재 진행 상황", ["완료", "진행중(이월)"])
        else:
            status = "예정"
            
        submit_button = st.form_submit_button("💾 시트에 기록하기")
        
        if submit_button:
            start_h, start_m = map(int, start_time_str.split(':'))
            end_h, end_m = map(int, end_time_str.split(':'))
            
            if (start_h > end_h) or (start_h == end_h and start_m >= end_m):
                st.error("종료 시간이 시작 시간보다 빠르거나 같을 수 없습니다!")
            elif not author_name.strip():
                st.warning("작성자 성명을 입력해주세요!")
            elif not project_name.strip():
                st.warning("고객사_프로젝트명을 입력해주세요!")
            else:
                new_data = pd.DataFrame([{
                    "날짜": date_input.strftime("%Y-%m-%d"),
                    "작성자": author_name.strip(),
                    "직급": rank,
                    "업무구분": auto_task_type,
                    "고객사_프로젝트명": project_name.strip(),
                    "시작시간": start_time_str,
                    "종료시간": end_time_str,
                    "업무량/내용": description,
                    "진행상황": status
                }])
                
                updated_df = pd.concat([df, new_data], ignore_index=True)
                conn.update(data=updated_df)
                st.success(f"🎉 데이터가 구글 시트에 실시간으로 누적되었습니다!")
                st.rerun()

# ----------------------------------------------------
# 오른쪽 화면: 일일 업무보고서 다운로드 엔진 (비즈니스 로직 맵핑 전면 수정)
# ----------------------------------------------------
with col2:
    st.header("📥 일일 업무보고서 다운로드")
    
    if df.empty or len(df.dropna(subset=["작성자"])) == 0:
        st.info("아직 기록된 팀 데이터가 없습니다.")
    else:
        df = df.sort_values(by="날짜").reset_index(drop=True)
        
        all_authors = sorted(df["작성자"].dropna().unique().tolist())
        selected_author = st.selectbox("📊 직원을 선택하세요", all_authors)
        
        author_df = df[df["작성자"] == selected_author].copy()
        
        # 🛠️ [중요] 날짜 매핑 로직 전환을 위해 데이터프레임 전처리 구현
        # 각 행마다 실제 배치되어야 할 'target_date'(보고서 대상 날짜)를 계산합니다.
        target_dates = []
        for _, row in author_df.iterrows():
            current_dt = datetime.datetime.strptime(row["날짜"], "%Y-%m-%d").date()
            if row["업무구분"] == "내일 계획":
                # '내일 계획'의 날짜가 6/24라면 실제 보고서는 6/23 시트에 작성되어야 하므로 하루를 뺍니다.
                target_dates.append((current_dt - datetime.timedelta(days=1)).strftime("%Y-%m-%d"))
            else:
                # '오늘 업무'는 그대로 유지합니다.
                target_dates.append(row["날짜"])
                
        author_df["target_date"] = target_dates
        author_df["연월"] = pd.to_datetime(author_df["target_date"]).dt.strftime("%Y-%m")
        
        all_months = sorted(author_df["연월"].unique().tolist(), reverse=True)
        selected_month = st.selectbox("📅 다운로드할 월 선택", all_months)
        
        # 선택한 월의 target_date 기준 데이터 필터링
        final_df = author_df[author_df["연월"] == selected_month]
        current_rank = final_df["직급"].iloc[-1] if not final_df.empty else "사원"
        
        st.write(f"👉 **{selected_author} {current_rank}**님의 서식이 준비되었습니다.")
        
        if st.button("📊 템플릿 서식으로 보고서 생성하기"):
            try:
                # 구글 시트 등록 날짜가 아닌, 실제 엑셀 시트 탭이 될 target_date 기준으로 고유 날짜를 추출합니다.
                unique_target_dates = sorted(final_df["target_date"].unique())
                output_buffer = io.BytesIO()
                final_wb = openpyxl.Workbook()
                default_sheet = final_wb.active
                
                for t_date in unique_target_dates:
                    template_wb = openpyxl.load_workbook("template.xlsx")
                    ws = template_wb.active
                    
                    sheet_title = pd.to_datetime(t_date).strftime("%m-%d")
                    # 해당 시트 날짜(t_date)에 들어가야 하는 데이터 수집
                    day_data = final_df[final_df["target_date"] == t_date]
                    
                    morning_tasks = []
                    afternoon_tasks = []
                    next_tasks = []
                    
                    for _, row in day_data.iterrows():
                        task_info = {
                            "project": row['고객사_프로젝트명'],
                            "content": f"{row['업무량/내용']}",
                            "time": f"{row['시작시간']} ~ {row['종료시간']}"
                        }
                        
                        # 💡 이미 target_date로 분류되어 들어왔으므로 업무구분 분기만 태우면 완벽합니다.
                        if row['업무구분'] == "내일 계획":
                            next_tasks.append(task_info)
                        else:
                            start_hour = int(row['시작시간'].split(':')[0])
                            if start_hour < 12:
                                morning_tasks.append(task_info)
                            else:
                                afternoon_tasks.append(task_info)
                    
                    # 📌 상단 고정 메타 정보 기입 (I열 적용)
                    ws["I2"] = t_date          
                    ws["I4"] = selected_author  
                    ws["I6"] = current_rank     
                    
                    # 1. 오전 업무 적재 (10번 행 고정 시작)
                    start_morning_row = 10
                    inserted_morning_count = 0
                    
                    for i, task in enumerate(morning_tasks):
                        current_r = start_morning_row + i
                        if i >= 1:
                            ws.insert_rows(current_r, 1)
                            inserted_morning_count += 1
                            for col_idx in range(1, max(ws.max_column, 24) + 1):
                                source_cell = ws.cell(row=current_r-1, column=col_idx)
                                new_cell = ws.cell(row=current_r, column=col_idx)
                                if source_cell.font:
                                    new_cell.font = Font(name=source_cell.font.name, size=source_cell.font.size)
                                if source_cell.border:
                                    new_cell.border = Border(left=source_cell.border.left, right=source_cell.border.right, top=source_cell.border.top, bottom=source_cell.border.bottom)
                                if source_cell.fill and source_cell.fill.fill_type:
                                    new_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type, start_color=source_cell.fill.start_color, end_color=source_cell.fill.end_color)
                                if source_cell.alignment:
                                    new_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal, vertical=source_cell.alignment.vertical)
                        
                        ws.cell(row=current_r, column=3, value=task["project"])   # C열
                        ws.cell(row=current_r, column=14, value=task["content"]) # N열
                        ws.cell(row=current_r, column=16, value=task["time"])    # P열

                    # 2. 오후 업무 적재 (오전 행 증가량에 연동)
                    start_afternoon_row = 12 + inserted_morning_count
                    inserted_afternoon_count = 0
                    
                    for i, task in enumerate(afternoon_tasks):
                        current_r = start_afternoon_row + i
                        if i >= 1:
                            ws.insert_rows(current_r, 1)
                            inserted_afternoon_count += 1
                            for col_idx in range(1, max(ws.max_column, 24) + 1):
                                source_cell = ws.cell(row=current_r-1, column=col_idx)
                                new_cell = ws.cell(row=current_r, column=col_idx)
                                if source_cell.font:
                                    new_cell.font = Font(name=source_cell.font.name, size=source_cell.font.size)
                                if source_cell.border:
                                    new_cell.border = Border(left=source_cell.border.left, right=source_cell.border.right, top=source_cell.border.top, bottom=source_cell.border.bottom)
                                if source_cell.fill and source_cell.fill.fill_type:
                                    new_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type, start_color=source_cell.fill.start_color, end_color=source_cell.fill.end_color)
                                if source_cell.alignment:
                                    new_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal, vertical=source_cell.alignment.vertical)
                        
                        ws.cell(row=current_r, column=3, value=task["project"])   # C열
                        ws.cell(row=current_r, column=14, value=task["content"]) # N열
                        ws.cell(row=current_r, column=16, value=task["time"])    # P열

                    # 3. 익일 업무 계획 적재 (동적 수식 연동 정밀 매핑)
                    start_plan_row = 17 + inserted_morning_count + inserted_afternoon_count
                    
                    for i, task in enumerate(next_tasks):
                        current_r = start_plan_row + i
                        if i >= 1:
                            ws.insert_rows(current_r, 1)
                            for col_idx in range(1, max(ws.max_column, 24) + 1):
                                source_cell = ws.cell(row=current_r-1, column=col_idx)
                                new_cell = ws.cell(row=current_r, column=col_idx)
                                if source_cell.font:
                                    new_cell.font = Font(name=source_cell.font.name, size=source_cell.font.size)
                                if source_cell.border:
                                    new_cell.border = Border(left=source_cell.border.left, right=source_cell.border.right, top=source_cell.border.top, bottom=source_cell.border.bottom)
                                if source_cell.fill and source_cell.fill.fill_type:
                                    new_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type, start_color=source_cell.fill.start_color, end_color=source_cell.fill.end_color)
                                if source_cell.alignment:
                                    new_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal, vertical=source_cell.alignment.vertical)
                        
                        ws.cell(row=current_r, column=2, value=f"{i+1:02d}")                    # B열 (순번)
                        ws.cell(row=current_r, column=3, value=task["project"])               # C열 (프로젝트)
                        ws.cell(row=current_r, column=14, value=task["content"])              # N열 (계획내용)

                    # 🛡️ 안전하게 최종 시트 병합 및 복사 처리
                    new_ws = final_wb.create_sheet(title=sheet_title)
                    
                    for col in ws.columns:
                        col_letter = openpyxl.utils.get_column_letter(col[0].column)
                        new_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
                    for r_idx in range(1, ws.max_row + 1):
                        if r_idx in ws.row_dimensions:
                            new_ws.row_dimensions[r_idx].height = ws.row_dimensions[r_idx].height
                    
                    for row in ws.iter_rows(values_only=False):
                        for cell in row:
                            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                            if cell.has_style:
                                new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                                new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
                                if cell.fill and cell.fill.fill_type:
                                    new_cell.fill = PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                                new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)
                                new_cell.number_format = cell.number_format
                    
                    for merged_range in ws.merged_cells.ranges:
                        new_ws.merge_cells(str(merged_range))
                
                if len(final_wb.sheetnames) > 1 and "Sheet" in final_wb.sheetnames:
                    final_wb.remove(default_sheet)
                    
                final_wb.save(output_buffer)
                st.success("✨ 날짜 조건 교차 반영 및 템플릿 빌드가 완료되었습니다! 아래 다운로드 버튼을 누르세요.")
                
                st.download_button(
                    label=f"📥 {selected_author}_일일업무보고서 양식 출력본 다운로드",
                    data=output_buffer.getvalue(),
                    file_name=f"{selected_month}_{selected_author}_{current_rank}_일일업무보고서.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as ex:
                import traceback
                st.error("🚨 에러가 발생한 위치와 상세 로그입니다:")
                st.code(traceback.format_exc(), language="python")
